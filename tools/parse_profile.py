#!/usr/bin/env python
import sys
import struct
import typing
import inspect
import datetime
import functools

import openpyxl

def noop(x): return x

def resolve(value, values): 
    return values.get(value, {}).get("name", value)


@functools.lru_cache(typed=True)
def picker(values, filters):
    try:
        return next(
            filter(
                lambda x: all(x[1][k] == v for k, v in filters.items()),
                values.items()
            )
        )
    except StopIteration:
        return (None, None)


def render(data):
    if isinstance(data, typing.List):
        out = ["list(["]
        for v in data:
            out.append(f"  {render(v)},")
        out.append("])")
        return "\n".join(out)
    elif isinstance(data, typing.Dict):
        out = ["dict({"]
        for k, v in data.items():
            out.append(f"  {render(k)}: {render(v)},")
        out.append("})")
        return "\n".join(out)
    elif isinstance(data, type(noop)):
        return data.__qualname__
    elif isinstance(data, str):
        return repr(data)
    elif isinstance(data, int):
        return str(data)
    elif isinstance(data, type(None)):
        return "None"
    return data


def parse_component_value(component, value):
    if value is None:
        return None

    if isinstance(value, tuple):
        if component["bit_offset"] and component["bit_offset"] >= len(value) << 3:
            raise ValueError()

        unpacked_num = 0
        for v in reversed(value):
            unpacked_num = (unpacked_num << 8) + v
        value = unpacked_num

    if isinstance(value, int):
        value = (value >> component["bit_offset"]) & ((1 << component["bits"]) - 1)

    return value



class dict(dict):
    def __hash__(self):
        return hash(tuple(self.items()))

class list(list):
    def __hash__(self):
        return hash(tuple(self))


class BaseTypeParsers():
    def enum(v, *_):
        return None if v == 0xFF else v

    def sint8(v, *_):
        return None if v == 0x7F else v

    def uint8(v, *_):
        return None if v == 0xFF else v

    def string(v, *_):
        if v is None:
            return None
        elif isinstance(v, bytes):
            v = v.decode('utf-8', errors="replace")
        try:
            s = v[:v.index('\x00')]
        except ValueError:
            s = v
        return s or None

    def uint8z(v, *_):
        return None if v == 0x0 else v

    def byte(v, *_):
        if v is None:
            return None
        if isinstance(v, int):
            return None if v == 0xFF else v
        return None if all(b == 0xFF for b in v) else v

    def sint16(v, *_):
        return None if v == 0x7FFF else v

    def uint16(v, *_):
        return None if v == 0xFFFF else v

    def sint32(v, *_):
        return None if v == 0x7FFFFFFF else v

    def uint32(v, *_):
        return None if v == 0xFFFFFFFF else v

    def float32(v, *_):
        return None if math.isnan(v) else v

    def float64(v, *_):
        return None if math.isnan(v) else v

    def uint16z(v, *_):
        return None if v == 0x0 else v

    def uint32z(v, *_):
        return None if v == 0x0 else v

    def sint64(v, *_):
        return None if v == 0x7FFFFFFFFFFFFFFF else v

    def uint64(v, *_):
        return None if v == 0xFFFFFFFFFFFFFFFF else v

    def uint64z(v, *_):
        return None if v == 0 else v

    def bool(v, *_):
        return True if v else False


class TypeParsers():
    def date_time(value, values):
        if value is None:
            return value
        if isinstance(value, datetime.datetime):
            return value
        max_offset_value = picker(values, dict(name="min") )[0]
        if value < max_offset_value:
            return value
        return datetime.datetime.utcfromtimestamp(value + GARMIN_EPOC_OFFSET) 
    local_date_time = date_time

    #def product(value, values):
    #    if value in values:
    #        return values[value]["name"]
    #    return value
    #product_name = product
    #garmin_product = product


base_type_list = [
    {"fmt":'B', "size":struct.calcsize('B'), "parser":BaseTypeParsers.enum,    "name":'enum'},
    {"fmt":'b', "size":struct.calcsize('b'), "parser":BaseTypeParsers.sint8,   "name":'sint8'},
    {"fmt":'B', "size":struct.calcsize('B'), "parser":BaseTypeParsers.uint8,   "name":'uint8'},
    {"fmt":'s', "size":struct.calcsize('s'), "parser":BaseTypeParsers.string,  "name":'string'},
    {"fmt":'B', "size":struct.calcsize('B'), "parser":BaseTypeParsers.uint8z,  "name":'uint8z'},
    {"fmt":'B', "size":struct.calcsize('B'), "parser":BaseTypeParsers.byte,    "name":'byte'},
    {"fmt":'h', "size":struct.calcsize('h'), "parser":BaseTypeParsers.sint16,  "name":'sint16'},
    {"fmt":'H', "size":struct.calcsize('H'), "parser":BaseTypeParsers.uint16,  "name":'uint16'},
    {"fmt":'i', "size":struct.calcsize('i'), "parser":BaseTypeParsers.sint32,  "name":'sint32'},
    {"fmt":'I', "size":struct.calcsize('I'), "parser":BaseTypeParsers.uint32,  "name":'uint32'},
    {"fmt":'f', "size":struct.calcsize('f'), "parser":BaseTypeParsers.float32, "name":'float32'},
    {"fmt":'d', "size":struct.calcsize('d'), "parser":BaseTypeParsers.float64, "name":'float64'},
    {"fmt":'H', "size":struct.calcsize('H'), "parser":BaseTypeParsers.uint16z, "name":'uint16z'},
    {"fmt":'I', "size":struct.calcsize('I'), "parser":BaseTypeParsers.uint32z, "name":'uint32z'},
    {"fmt":'q', "size":struct.calcsize('q'), "parser":BaseTypeParsers.sint64,  "name":'sint64'},
    {"fmt":'Q', "size":struct.calcsize('Q'), "parser":BaseTypeParsers.uint64,  "name":'uint64'},
    {"fmt":'Q', "size":struct.calcsize('Q'), "parser":BaseTypeParsers.uint64z, "name":'uint64z'},
#    {"fmt":'B', "size":struct.calcsize('B'), "parser":BaseTypeParsers.bool,    "name":'bool'},
]


def parse_csv_field(field, items=0, func=lambda x: x):
    if not field:
        return [func(None)] * items
    if isinstance(field, (str, bytes)):
        array = [func(int(striped)) if striped.isdigit() else func(striped) for raw in field.split(',') if (striped := raw.strip())]
    else:
        array = [func(field)]

    array_lenght = len(array)
    if array_lenght < items:
        assert array_lenght == 1
        return array * items
    return array

fix_scale = lambda x: None if x == 1 else x
fix_units = lambda x: x.replace(' / ', '/').replace(' * ', '*').replace('(steps)', 'or steps') if x else None

#
## TYPES
#
types = dict()
def parse_types_sheet(sheet):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        type_name, base_type, value_name, value, comment = row
        if comment:
            try:
                comment = comment.decode("utf-8", ignore="replace").strip()
            except AttributeError:
                comment = comment.strip()
        else:
            comment = None

        if type_name:
            type_parser = getattr(TypeParsers, type_name, resolve)
            current_type = types[type_name] = dict({
                "name":type_name,
                "base_type": base_type, 
                "parser":type_parser,
                "values": {}
            })
            if comment:
                current_type["comment"] = comment
            
        elif value is not None:
            try:
                int_value = int(value)
            except ValueError as e:
                int_value = int(value, 16)

            current_type["values"][int_value] = dict({
                "name":value_name,
                "value":int_value
            })

            if not comment:
                continue

            current_type["values"][int_value]["comment"] = comment

            if not comment.startswith(f"0x{hex(int_value).upper()[2:]}"):  # no range values
                continue

            start = int_value + 1
            end = int(comment.split()[2], 16)
            name = "_".join(value_name.split("_")[:-1])
            for n in range(start, end):
                current_type["values"][n] = dict({
                    "name": f"{name}_{n}",
                    "value": n,
                })
                if comment:
                    current_type["values"][n]["comment"] = comment

        else:
            assert not any(row), "We should not get here"


base_types = {}
def format_base_types():
    base_type_map = {v["name"]:k for k, v in types['fit_base_type']["values"].items()}
    base_type_map["bool"] = 0
    for base_type_kwargs in base_type_list: 
        byte = base_type_map.get(base_type_kwargs["name"])
        base_types[base_type_kwargs["name"]] = dict({**{**base_type_kwargs, **{"byte": byte, "type_num": byte & 0x1F}}})


def get_type(type):
    if not isinstance(type, str):
        return type
    return base_types.get(type) or types.get(type)


#
## MESSAGES
#
messages = {}
def parse_message_sheet(sheet):
    global messages
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(bool(v) for i, v in enumerate(row) if i != 3):  # empty line or message group header
            if row[3] and len(row[3]):  # message group header
                group_name = row[3].title()
            continue
        (
            message_name, field_def_num, field_name, field_type, array,
            components, scale, offset, units, bits, accumulate,
            ref_field_name, ref_field_value, comment, products, example
        ) = row[:-3]

        if message_name:
            global_number = next(filter(lambda x: x[1]["name"] == message_name, types["mesg_num"]["values"].items()))[0]
            current_message = messages[global_number] = dict(
                    name=message_name,
                    global_number=global_number,
                    group_name=group_name,
                    fields=dict(),
                    comment=comment
            )
            continue

        component_fields = list()
        if components and (component_names := [c.strip() for c in components.split(",") if c.strip()]):

            num_components = len(component_names)

            component_fields.extend(
                list(
                    dict(
                        name=component_name,
                        scale=component_scale,
                        offset=component_offset,
                        units=component_units,
                        bits=component_bits,
                        accumulate=component_accumulate
                    ) 
                    for component_name, component_scale, component_offset, component_units, component_bits, component_accumulate in zip(
                        component_names,
                        parse_csv_field(scale, num_components, fix_scale),
                        parse_csv_field(offset, num_components),
                        parse_csv_field(units, num_components, fix_units),
                        parse_csv_field(bits, num_components),
                        parse_csv_field(accumulate, num_components, bool)
                    )
                )
            )
            
            assert len(component_fields) == num_components
            assert all(c["name"] and c["bits"] for c in component_fields)

        if field_def_num is not None:
            current_field = current_message["fields"][field_def_num] = dict(
                name=field_name,
                type=field_type,
                def_num=field_def_num,
                scale=fix_scale(scale),
                offset=offset,
                units=fix_units(units),
                subfields=list(),
                components=component_fields
            )
            if comment:
                current_field["comment"] = comment
            continue

        elif field_name is not None:
            subfield = dict(
                name=field_name,
                type=field_type,
                scale=fix_scale(scale),
                offset=offset,
                units=fix_units(units),
                reference_fields=[],
                components=component_fields,

            )
            if comment:
                subfield["comment"] = comment

            ref_field_names = parse_csv_field(ref_field_name)
            assert ref_field_names, f"{ref_field_name} {ref_field_names} {row}"

            subfield["reference_fields"].extend(
                list(
                    dict(
                        name=name,
                        value=value,
                        def_num=None,
                        raw_value=None
                    ) for name, value in zip(ref_field_names, parse_csv_field(ref_field_value))
                )
            )

            assert len(subfield["reference_fields"]) == len(ref_field_names)
            if "alert_type" not in ref_field_names:
                current_field["subfields"].append(subfield)

            continue

    messages = dict({k:v for k ,v in sorted(messages.items(), key=lambda item: item[1]["global_number"])})

    for message_name, message in messages.items():
        for _, field in message["fields"].items():
            for subfield in field["subfields"]:
                for sub_ref_field in subfield["reference_fields"]:
                    ref_field = next(filter(lambda x: x[1]["name"] == sub_ref_field["name"], message["fields"].items()))[1]
                    sub_ref_field["def_num"] = ref_field["def_num"] 
                    sub_ref_field["raw_value"] = next(
                        filter(
                            lambda x: x[1]["name"] == sub_ref_field["value"],
                            get_type(ref_field["type"])["values"].items()
                        )
                    )[0] 

                bit_offset = 0
                for component in subfield["components"]:
                    component["num"] = next(
                        filter(
                            lambda x: x[1]["name"] == component["name"],
                            message["fields"].items()
                        )
                    )[0]
                    component["bit_offset"] = bit_offset
                    bit_offset += component["bits"]
            bit_offset = 0
            for component in field["components"]:
                component["num"] = next(
                    filter(
                        lambda x: x[1]["name"] == component["name"],
                        message["fields"].items()
                    )
                )[0]
                component["bit_offset"] = bit_offset
                bit_offset += component["bits"]


GARMIN_EPOC_OFFSET = datetime.datetime.fromisoformat('1989-12-31T00:00:00+00:00').timestamp()

obj2render = [
    noop,
    picker,
    resolve,
    parse_component_value,
    dict,
    list,
    BaseTypeParsers,
    TypeParsers
]


def print_rendered():
    print("# -*- coding: utf-8 -*-\n")
    print("import math")
    print("import datetime")
    print("import functools")
    print("")
    print(f"{GARMIN_EPOC_OFFSET=}")
    print("")
    for obj in obj2render:
        print(inspect.getsource(obj))
        print("\n")
    print(f"base_types = {render(base_types)}")
    print(f"types = {render(types)}")
    print(f"messages = {render(messages)}")


def dofile(file):
    wb = openpyxl.load_workbook(file)
    ws_types = wb['Types']
    ws_messages = wb['Messages']
    parse_types_sheet(ws_types)
    format_base_types()
    parse_message_sheet(ws_messages)
    print_rendered()


if __name__ == "__main__":
    dofile(sys.argv[-1])
